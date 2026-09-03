from io import BytesIO
from fastapi.responses import StreamingResponse
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.issue import Issue
from app.models.book import Book
from app.models.user import User
from app.utils.dependencies import require_roles

#from datetime import datetime
# from sqlalchemy import func
from sqlalchemy import func, case


from app.models.reservation import Reservation
from app.models.user import User
from app.models.book import Book

from app.models.role import Role
from datetime import date, datetime, time
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


@router.get("/issued-books")
def get_issued_books_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):

    query = (
        db.query(Issue, User, Book)
        .join(User, Issue.user_id == User.id)
        .join(Book, Issue.book_id == Book.id)
        .filter(Issue.status == "ISSUED")
    )

    if start_date:
        start_datetime = datetime.combine(
            start_date,
            time.min
        )

        query = query.filter(
            Issue.issue_date >= start_datetime
        )

    if end_date:
        end_datetime = datetime.combine(
            end_date,
            time.max
        )

        query = query.filter(
            Issue.issue_date <= end_datetime
        )

    issued_records = (
        query
        .order_by(Issue.issue_date.desc())
        .all()
    )

    report = []

    for issue, user, book in issued_records:

        report.append({
            "issue_id": issue.id,

            "member": {
                "user_id": user.id,
                "username": user.username,
                "full_name": user.full_name
            },

            "book": {
                "book_id": book.id,
                "title": book.title,
                "isbn": book.isbn
            },

            "issue_date": _excel_safe_datetime(issue.issue_date),
            "due_date": _excel_safe_datetime(issue.due_date),
            "status": issue.status
        })

    return {
        "start_date": start_date,
        "end_date": end_date,
        "total_issued_books": len(report),
        "issued_books": report
    }






@router.get("/overdue-books")
def get_overdue_books_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):

    now = datetime.utcnow()

    overdue_records = (
        db.query(Issue, User, Book)
        .join(User, Issue.user_id == User.id)
        .join(Book, Issue.book_id == Book.id)
        .filter(
            Issue.status == "ISSUED",
            Issue.due_date < now
        )
        .order_by(Issue.due_date.asc())
        .all()
    )

    report = []

    for issue, user, book in overdue_records:

        overdue_days = (
            now.date() - issue.due_date.date()
        ).days

        current_fine = overdue_days * 5

        report.append({
            "issue_id": issue.id,

            "member": {
                "user_id": user.id,
                "username": user.username,
                "full_name": user.full_name
            },

            "book": {
                "book_id": book.id,
                "title": book.title,
                "isbn": book.isbn
            },

            "issue_date": _excel_safe_datetime(issue.issue_date),
            "due_date": _excel_safe_datetime(issue.due_date),
            "overdue_days": overdue_days,
            "current_fine": current_fine,
            "status": issue.status
        })

    return {
        "total_overdue_books": len(report),
        "overdue_books": report
    }


@router.get("/fines")
def get_fine_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):

    fine_records = (
        db.query(Issue, User, Book)
        .join(User, Issue.user_id == User.id)
        .join(Book, Issue.book_id == Book.id)
        .filter(Issue.fine_amount > 0)
        .order_by(Issue.id.desc())
        .all()
    )

    report = []

    total_generated = 0
    total_paid = 0
    total_outstanding = 0

    for issue, user, book in fine_records:

        fine_amount = float(issue.fine_amount or 0)

        total_generated += fine_amount

        if issue.fine_status == "PAID":
            total_paid += fine_amount
        else:
            total_outstanding += fine_amount

        report.append({
            "issue_id": issue.id,
            "user_id": user.id,
            "username": user.username,
            "book_id": book.id,
            "book_title": book.title,
            "fine_amount": fine_amount,
            "fine_status": issue.fine_status,
            "fine_paid_at": issue.fine_paid_at
        })

    return {
        "total_fine_cases": len(report),
        "total_fines_generated": total_generated,
        "total_fines_paid": total_paid,
        "total_fines_outstanding": total_outstanding,
        "fine_records": report
    }



@router.get("/most-borrowed-books")
def get_most_borrowed_books_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):
    results = (
        db.query(
            Book.id.label("book_id"),
            Book.title.label("title"),
            Book.isbn.label("isbn"),
            func.count(Issue.id).label("borrow_count")
        )
        .join(Issue, Issue.book_id == Book.id)
        .group_by(
            Book.id,
            Book.title,
            Book.isbn
        )
        .order_by(
            func.count(Issue.id).desc(),
            Book.title.asc()
        )
        .all()
    )

    report = []

    for row in results:
        report.append({
            "book_id": row.book_id,
            "title": row.title,
            "isbn": row.isbn,
            "borrow_count": row.borrow_count
        })

    return {
        "total_books": len(report),
        "most_borrowed_books": report
    }



@router.get("/reservations")
def get_reservation_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):
    reservations = (
        db.query(
            Reservation,
            User.username,
            Book.title
        )
        .join(User, User.id == Reservation.user_id)
        .join(Book, Book.id == Reservation.book_id)
        .order_by(Reservation.id.desc())
        .all()
    )

    records = []

    for reservation, username, book_title in reservations:
        records.append({
            "reservation_id": reservation.id,
            "user_id": reservation.user_id,
            "username": username,
            "book_id": reservation.book_id,
            "book_title": book_title,
            "reserved_at": _excel_safe_datetime(reservation.reserved_at),
            "status": reservation.status,
            "ready_until": reservation.ready_until
        })

    return {
        "total_reservations": len(records),
        "reservation_records": records
    }



@router.get("/member-activity")
def get_member_activity_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):
    results = (
        db.query(
            User.id.label("user_id"),
            User.username.label("username"),
            User.full_name.label("full_name"),

            func.count(Issue.id).label("total_borrowed"),

            func.sum(
                case(
                    (Issue.status == "ISSUED", 1),
                    else_=0
                )
            ).label("currently_issued"),

            func.sum(
                case(
                    (Issue.status == "RETURNED", 1),
                    else_=0
                )
            ).label("returned_books"),

            func.coalesce(
                func.sum(Issue.fine_amount),
                0
            ).label("total_fines")
        )
        .outerjoin(
            Issue,
            Issue.user_id == User.id
        )
        .join(
            Role,
            User.role_id == Role.id
        )
        .filter(
            Role.name == "MEMBER"
        )
        .group_by(
            User.id,
            User.username,
            User.full_name
        )
        .order_by(
            func.count(Issue.id).desc()
        )
        .all()
    )

    report = []

    for row in results:
        report.append({
            "user_id": row.user_id,
            "username": row.username,
            "full_name": row.full_name,
            "total_borrowed": row.total_borrowed or 0,
            "currently_issued": row.currently_issued or 0,
            "returned_books": row.returned_books or 0,
            "total_fines": float(row.total_fines or 0)
        })

    return {
        "total_users": len(report),
        "member_activity": report
    }


@router.get("/issue-history")
def get_issue_history_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):

    query = (
        db.query(Issue, User, Book)
        .join(User, Issue.user_id == User.id)
        .join(Book, Issue.book_id == Book.id)
    )

    if start_date:
        start_datetime = datetime.combine(
            start_date,
            time.min
        )

        query = query.filter(
            Issue.issue_date >= start_datetime
        )

    if end_date:
        end_datetime = datetime.combine(
            end_date,
            time.max
        )

        query = query.filter(
            Issue.issue_date <= end_datetime
        )

    issue_records = (
        query
        .order_by(Issue.issue_date.desc())
        .all()
    )

    report = []

    for issue, user, book in issue_records:

        report.append({
            "issue_id": issue.id,

            "member": {
                "user_id": user.id,
                "username": user.username,
                "full_name": user.full_name
            },

            "book": {
                "book_id": book.id,
                "title": book.title,
                "isbn": book.isbn
            },

            "issue_date": _excel_safe_datetime(issue.issue_date),
            "due_date": _excel_safe_datetime(issue.due_date),
            "return_date": _excel_safe_datetime(issue.return_date),
            "status": issue.status,
            "overdue_days": issue.overdue_days,
            "fine_amount": float(issue.fine_amount or 0),
            "fine_status": issue.fine_status,
            "renewal_count": issue.renewal_count
        })

    return {
        "start_date": start_date,
        "end_date": end_date,
        "total_issue_records": len(report),
        "issue_history": report
    }


@router.get("/summary")
def get_reports_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    )
):

    total_members = (
        db.query(func.count(User.id))
        .join(Role, User.role_id == Role.id)
        .filter(Role.name == "MEMBER")
        .scalar()
    ) or 0

    total_books = (
        db.query(func.count(Book.id))
        .scalar()
    ) or 0

    currently_issued = (
        db.query(func.count(Issue.id))
        .filter(Issue.status == "ISSUED")
        .scalar()
    ) or 0

    returned_books = (
        db.query(func.count(Issue.id))
        .filter(Issue.status == "RETURNED")
        .scalar()
    ) or 0

    active_reservations = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.status == "ACTIVE")
        .scalar()
    ) or 0

    overdue_books = (
        db.query(func.count(Issue.id))
        .filter(
            Issue.status == "ISSUED",
            Issue.due_date < datetime.utcnow()
        )
        .scalar()
    ) or 0

    total_fines_generated = (
        db.query(
            func.coalesce(
                func.sum(Issue.fine_amount),
                0
            )
        )
        .filter(Issue.fine_amount > 0)
        .scalar()
    ) or 0

    total_fines_paid = (
        db.query(
            func.coalesce(
                func.sum(Issue.fine_amount),
                0
            )
        )
        .filter(
            Issue.fine_amount > 0,
            Issue.fine_status == "PAID"
        )
        .scalar()
    ) or 0

    total_fines_outstanding = (
        db.query(
            func.coalesce(
                func.sum(Issue.fine_amount),
                0
            )
        )
        .filter(
            Issue.fine_amount > 0,
            Issue.fine_status != "PAID"
        )
        .scalar()
    ) or 0

    return {
        "total_members": total_members,
        "total_books": total_books,
        "currently_issued": currently_issued,
        "returned_books": returned_books,
        "active_reservations": active_reservations,
        "overdue_books": overdue_books,
        "total_fines_generated": float(total_fines_generated),
        "total_fines_paid": float(total_fines_paid),
        "total_fines_outstanding": float(total_fines_outstanding)
    }


def _excel_safe_datetime(value):
    """
    openpyxl cannot write timezone-aware datetime/time objects.
    Convert them to timezone-naive values before writing to Excel.
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value

    if isinstance(value, time):
        if value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value

    return value


# --------------------------------------------------
# EXECUTIVE MANAGEMENT REPORT EXPORTS
# ADMIN ONLY
# --------------------------------------------------

def _executive_report_data(db: Session):
    now = datetime.utcnow()

    total_members = (
        db.query(func.count(User.id))
        .join(Role, User.role_id == Role.id)
        .filter(Role.name == "MEMBER")
        .scalar()
    ) or 0

    active_titles = (
        db.query(func.count(Book.id))
        .filter(Book.is_active == True)
        .scalar()
    ) or 0

    total_copies = (
        db.query(
            func.coalesce(func.sum(Book.total_copies), 0)
        )
        .filter(Book.is_active == True)
        .scalar()
    ) or 0

    available_copies = (
        db.query(
            func.coalesce(func.sum(Book.available_copies), 0)
        )
        .filter(Book.is_active == True)
        .scalar()
    ) or 0

    total_issues = db.query(func.count(Issue.id)).scalar() or 0

    returned_issues = (
        db.query(func.count(Issue.id))
        .filter(Issue.status == "RETURNED")
        .scalar()
    ) or 0

    currently_issued = (
        db.query(func.count(Issue.id))
        .filter(Issue.status == "ISSUED")
        .scalar()
    ) or 0

    overdue_books = (
        db.query(func.count(Issue.id))
        .filter(
            Issue.status == "ISSUED",
            Issue.due_date < now
        )
        .scalar()
    ) or 0

    active_reservations = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.status == "ACTIVE")
        .scalar()
    ) or 0

    ready_for_pickup = (
        db.query(func.count(Reservation.id))
        .filter(Reservation.status == "READY")
        .scalar()
    ) or 0

    fines_generated = float(
        db.query(
            func.coalesce(func.sum(Issue.fine_amount), 0)
        )
        .filter(Issue.fine_amount > 0)
        .scalar()
        or 0
    )

    fines_collected = float(
        db.query(
            func.coalesce(func.sum(Issue.fine_amount), 0)
        )
        .filter(
            Issue.fine_amount > 0,
            Issue.fine_status == "PAID"
        )
        .scalar()
        or 0
    )

    outstanding_fines = float(
        db.query(
            func.coalesce(func.sum(Issue.fine_amount), 0)
        )
        .filter(
            Issue.fine_amount > 0,
            Issue.fine_status != "PAID"
        )
        .scalar()
        or 0
    )

    return_rate = (
        round((returned_issues / total_issues) * 100, 2)
        if total_issues > 0 else 0.0
    )

    fine_collection_rate = (
        round((fines_collected / fines_generated) * 100, 2)
        if fines_generated > 0 else 100.0
    )

    low_stock_titles = (
        db.query(func.count(Book.id))
        .filter(
            Book.is_active == True,
            Book.available_copies > 0,
            Book.available_copies <= 2
        )
        .scalar()
    ) or 0

    out_of_stock_titles = (
        db.query(func.count(Book.id))
        .filter(
            Book.is_active == True,
            Book.available_copies <= 0
        )
        .scalar()
    ) or 0

    locked_accounts = (
        db.query(func.count(User.id))
        .filter(
            User.locked_until.isnot(None),
            User.locked_until > now
        )
        .scalar()
    ) or 0

    priority_actions = []

    if overdue_books > 0:
        priority_actions.append(
            ("HIGH", "Circulation",
             f"Follow up on {overdue_books} overdue book(s).")
        )

    if outstanding_fines > 0:
        priority_actions.append(
            ("MEDIUM", "Fines",
             f"Rs. {outstanding_fines:.2f} remains outstanding.")
        )

    if out_of_stock_titles > 0:
        priority_actions.append(
            ("HIGH", "Inventory",
             f"{out_of_stock_titles} active title(s) are out of stock.")
        )

    if low_stock_titles > 0:
        priority_actions.append(
            ("MEDIUM", "Inventory",
             f"{low_stock_titles} active title(s) have low stock.")
        )

    if locked_accounts > 0:
        priority_actions.append(
            ("HIGH", "Security",
             f"{locked_accounts} account(s) are currently locked.")
        )

    if not priority_actions:
        priority_actions.append(
            ("LOW", "Operations",
             "No immediate executive intervention is required.")
        )

    if (
        locked_accounts > 0
        or out_of_stock_titles > 0
        or (currently_issued > 0 and overdue_books / currently_issued > 0.10)
    ):
        overall_status = "CRITICAL"
    elif (
        overdue_books > 0
        or outstanding_fines > 0
        or low_stock_titles > 0
        or active_reservations > 0
    ):
        overall_status = "ATTENTION"
    else:
        overall_status = "GOOD"

    metrics = [
        ("Overall Status", overall_status),
        ("Total Members", total_members),
        ("Active Book Titles", active_titles),
        ("Total Copies", int(total_copies)),
        ("Available Copies", int(available_copies)),
        ("Total Issue Records", total_issues),
        ("Returned Issues", returned_issues),
        ("Currently Issued", currently_issued),
        ("Overdue Books", overdue_books),
        ("Return Rate", f"{return_rate:.2f}%"),
        ("Active Reservations", active_reservations),
        ("Ready for Pickup", ready_for_pickup),
        ("Fines Generated", f"Rs. {fines_generated:.2f}"),
        ("Fines Collected", f"Rs. {fines_collected:.2f}"),
        ("Outstanding Fines", f"Rs. {outstanding_fines:.2f}"),
        ("Fine Collection Rate", f"{fine_collection_rate:.2f}%"),
        ("Low Stock Titles", low_stock_titles),
        ("Out of Stock Titles", out_of_stock_titles),
        ("Currently Locked Accounts", locked_accounts),
    ]

    return {
        "generated_at": now,
        "overall_status": overall_status,
        "metrics": metrics,
        "priority_actions": priority_actions,
    }


@router.get("/executive/pdf")
def export_executive_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN")
    )
):
    data = _executive_report_data(db)
    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Smart Library Executive Management Report",
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(
        Paragraph(
            "Smart Library Management System",
            styles["Title"]
        )
    )
    story.append(
        Paragraph(
            "Executive Management Report",
            styles["Heading2"]
        )
    )
    story.append(
        Paragraph(
            "Generated: "
            + data["generated_at"].strftime("%d-%m-%Y %H:%M:%S"),
            styles["Normal"]
        )
    )
    story.append(Spacer(1, 8))

    story.append(
        Paragraph(
            f"Overall Status: {data['overall_status']}",
            styles["Heading2"]
        )
    )
    story.append(Spacer(1, 6))

    metric_rows = [["KPI", "Value"]]
    metric_rows.extend(
        [[str(label), str(value)] for label, value in data["metrics"]]
    )

    metric_table = Table(
        metric_rows,
        colWidths=[95 * mm, 70 * mm],
        repeatRows=1,
    )
    metric_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
    )
    story.append(metric_table)
    story.append(Spacer(1, 14))

    story.append(
        Paragraph(
            "Management Priorities",
            styles["Heading2"]
        )
    )

    priority_rows = [["Priority", "Area", "Recommended Action"]]
    priority_rows.extend(
        [
            [priority, area, message]
            for priority, area, message
            in data["priority_actions"]
        ]
    )

    priority_table = Table(
        priority_rows,
        colWidths=[30 * mm, 45 * mm, 170 * mm],
        repeatRows=1,
    )
    priority_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
    )
    story.append(priority_table)

    document.build(story)
    buffer.seek(0)

    filename = (
        "smart_library_executive_report_"
        + data["generated_at"].strftime("%Y%m%d_%H%M%S")
        + ".pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )


@router.get("/executive/excel")
def export_executive_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN")
    )
):
    data = _executive_report_data(db)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Executive Summary"

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )
    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    ws["A1"] = "Smart Library Management System"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:B1")

    ws["A2"] = "Executive Management Report"
    ws["A2"].font = Font(bold=True, size=13)
    ws.merge_cells("A2:B2")

    ws["A3"] = "Generated At"
    ws["B3"] = data["generated_at"].strftime(
        "%d-%m-%Y %H:%M:%S"
    )

    ws["A4"] = "Overall Status"
    ws["B4"] = data["overall_status"]
    ws["B4"].font = Font(bold=True)

    ws.append([])
    ws.append(["KPI", "Value"])

    for cell in ws[6]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for label, value in data["metrics"]:
        ws.append([label, value])

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24

    priorities = workbook.create_sheet(
        "Management Priorities"
    )
    priorities.append(
        ["Priority", "Area", "Recommended Action"]
    )

    for cell in priorities[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for priority, area, message in data["priority_actions"]:
        priorities.append(
            [priority, area, message]
        )

    priorities.column_dimensions["A"].width = 14
    priorities.column_dimensions["B"].width = 24
    priorities.column_dimensions["C"].width = 70

    # Add detailed management worksheets.
    issue_sheet = workbook.create_sheet("Circulation")
    issue_sheet.append(
        [
            "Issue ID",
            "User ID",
            "Book ID",
            "Issue Date",
            "Due Date",
            "Return Date",
            "Status",
            "Fine Amount",
            "Fine Status",
            "Renewals",
        ]
    )

    for cell in issue_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    issue_rows = (
        db.query(Issue)
        .order_by(Issue.issue_date.desc())
        .all()
    )

    for issue in issue_rows:
        issue_sheet.append([
            issue.id,
            issue.user_id,
            issue.book_id,
            _excel_safe_datetime(issue.issue_date),
            _excel_safe_datetime(issue.due_date),
            _excel_safe_datetime(issue.return_date),
            issue.status,
            float(issue.fine_amount or 0),
            issue.fine_status,
            issue.renewal_count or 0,
        ])

    reservation_sheet = workbook.create_sheet(
        "Reservations"
    )
    reservation_sheet.append(
        [
            "Reservation ID",
            "User ID",
            "Book ID",
            "Reserved At",
            "Status",
            "Ready Until",
        ]
    )

    for cell in reservation_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    reservation_rows = (
        db.query(Reservation)
        .order_by(Reservation.id.desc())
        .all()
    )

    for reservation in reservation_rows:
        reservation_sheet.append([
            reservation.id,
            reservation.user_id,
            reservation.book_id,
            _excel_safe_datetime(reservation.reserved_at),
            reservation.status,
            _excel_safe_datetime(reservation.ready_until),
        ])

    inventory_sheet = workbook.create_sheet(
        "Inventory"
    )
    inventory_sheet.append(
        [
            "Book ID",
            "Title",
            "ISBN",
            "Total Copies",
            "Available Copies",
            "Active",
        ]
    )

    for cell in inventory_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    book_rows = (
        db.query(Book)
        .order_by(Book.title.asc())
        .all()
    )

    for book in book_rows:
        inventory_sheet.append([
            book.id,
            book.title,
            book.isbn,
            book.total_copies,
            book.available_copies,
            bool(book.is_active),
        ])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = (
            "A7"
            if sheet.title == "Executive Summary"
            else "A2"
        )

        for column_cells in sheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )

            if sheet.column_dimensions[column_letter].width is None:
                sheet.column_dimensions[column_letter].width = 18

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = (
        "smart_library_executive_report_"
        + data["generated_at"].strftime("%Y%m%d_%H%M%S")
        + ".xlsx"
    )

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )

# --------------------------------------------------
# ALL OPERATIONAL REPORT EXPORTS
# ADMIN / LIBRARIAN
# --------------------------------------------------

def _report_export_payload(
    report_name: str,
    db: Session,
    current_user: User,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    if report_name == "summary":
        data = get_reports_summary(
            db=db,
            current_user=current_user,
        )
        rows = [
            ["Total Members", data["total_members"]],
            ["Book Titles", data["total_books"]],
            ["Currently Issued", data["currently_issued"]],
            ["Returned Books", data["returned_books"]],
            ["Active Reservations", data["active_reservations"]],
            ["Overdue Books", data["overdue_books"]],
            ["Fines Generated", f"Rs. {data['total_fines_generated']:.2f}"],
            ["Fines Paid", f"Rs. {data['total_fines_paid']:.2f}"],
            ["Outstanding Fines", f"Rs. {data['total_fines_outstanding']:.2f}"],
        ]
        return {
            "title": "Summary Report",
            "headers": ["Metric", "Value"],
            "rows": rows,
        }

    if report_name == "issued-books":
        data = get_issued_books_report(
            start_date=start_date,
            end_date=end_date,
            db=db,
            current_user=current_user,
        )
        rows = []
        for row in data["issued_books"]:
            rows.append([
                row["issue_id"],
                row["member"].get("full_name")
                or row["member"].get("username")
                or f"User {row['member'].get('user_id')}",
                row["book"].get("title") or "-",
                row["book"].get("isbn") or "-",
                row.get("issue_date"),
                row.get("due_date"),
                row.get("status") or "-",
            ])
        return {
            "title": "Issued Books Report",
            "headers": [
                "Issue ID", "Member", "Book", "ISBN",
                "Issue Date", "Due Date", "Status"
            ],
            "rows": rows,
        }

    if report_name == "overdue-books":
        data = get_overdue_books_report(
            db=db,
            current_user=current_user,
        )
        rows = []
        for row in data["overdue_books"]:
            rows.append([
                row["issue_id"],
                row["member"].get("full_name")
                or row["member"].get("username")
                or "-",
                row["book"].get("title") or "-",
                row.get("due_date"),
                row.get("overdue_days", 0),
                f"Rs. {float(row.get('current_fine') or 0):.2f}",
                row.get("status") or "-",
            ])
        return {
            "title": "Overdue Books Report",
            "headers": [
                "Issue ID", "Member", "Book", "Due Date",
                "Overdue Days", "Current Fine", "Status"
            ],
            "rows": rows,
        }

    if report_name == "fines":
        data = get_fine_report(
            db=db,
            current_user=current_user,
        )
        rows = []
        for row in data["fine_records"]:
            rows.append([
                row["issue_id"],
                row.get("username") or "-",
                row.get("book_title") or "-",
                f"Rs. {float(row.get('fine_amount') or 0):.2f}",
                row.get("fine_status") or "-",
                _excel_safe_datetime(row.get("fine_paid_at")),
            ])
        return {
            "title": "Fines Report",
            "headers": [
                "Issue ID", "Member", "Book",
                "Fine Amount", "Fine Status", "Paid At"
            ],
            "rows": rows,
            "summary": [
                ["Fine Cases", data["total_fine_cases"]],
                ["Fines Generated", f"Rs. {data['total_fines_generated']:.2f}"],
                ["Fines Paid", f"Rs. {data['total_fines_paid']:.2f}"],
                ["Outstanding", f"Rs. {data['total_fines_outstanding']:.2f}"],
            ],
        }

    if report_name == "most-borrowed-books":
        data = get_most_borrowed_books_report(
            db=db,
            current_user=current_user,
        )
        rows = [
            [
                row["book_id"],
                row.get("title") or "-",
                row.get("isbn") or "-",
                row.get("borrow_count", 0),
            ]
            for row in data["most_borrowed_books"]
        ]
        return {
            "title": "Most Borrowed Books Report",
            "headers": [
                "Book ID", "Title", "ISBN", "Times Borrowed"
            ],
            "rows": rows,
        }

    if report_name == "reservations":
        data = get_reservation_report(
            db=db,
            current_user=current_user,
        )
        rows = []
        for row in data["reservation_records"]:
            rows.append([
                row["reservation_id"],
                row.get("username") or "-",
                row.get("book_title") or "-",
                _excel_safe_datetime(row.get("reserved_at")),
                row.get("status") or "-",
                _excel_safe_datetime(row.get("ready_until")),
            ])
        return {
            "title": "Reservations Report",
            "headers": [
                "Reservation ID", "Member", "Book",
                "Reserved At", "Status", "Ready Until"
            ],
            "rows": rows,
        }

    if report_name == "member-activity":
        data = get_member_activity_report(
            db=db,
            current_user=current_user,
        )
        rows = [
            [
                row["user_id"],
                row.get("full_name")
                or row.get("username")
                or "-",
                row.get("total_borrowed", 0),
                row.get("currently_issued", 0),
                row.get("returned_books", 0),
                f"Rs. {float(row.get('total_fines') or 0):.2f}",
            ]
            for row in data["member_activity"]
        ]
        return {
            "title": "Member Activity Report",
            "headers": [
                "User ID", "Member", "Total Borrowed",
                "Currently Issued", "Returned", "Total Fines"
            ],
            "rows": rows,
        }

    if report_name == "issue-history":
        data = get_issue_history_report(
            start_date=start_date,
            end_date=end_date,
            db=db,
            current_user=current_user,
        )
        rows = []
        for row in data["issue_history"]:
            rows.append([
                row["issue_id"],
                row["member"].get("full_name")
                or row["member"].get("username")
                or "-",
                row["book"].get("title") or "-",
                row.get("issue_date"),
                row.get("due_date"),
                row.get("return_date"),
                row.get("status") or "-",
                row.get("overdue_days") or 0,
                f"Rs. {float(row.get('fine_amount') or 0):.2f}",
                row.get("fine_status") or "-",
                row.get("renewal_count") or 0,
            ])
        return {
            "title": "Issue History Report",
            "headers": [
                "Issue ID", "Member", "Book", "Issued",
                "Due Date", "Returned", "Status",
                "Overdue Days", "Fine", "Fine Status",
                "Renewals"
            ],
            "rows": rows,
        }

    return None


def _report_date_filter_text(
    start_date: Optional[date],
    end_date: Optional[date],
):
    if not start_date and not end_date:
        return None

    start_text = (
        start_date.strftime("%d-%m-%Y")
        if start_date else "Beginning"
    )
    end_text = (
        end_date.strftime("%d-%m-%Y")
        if end_date else "Present"
    )
    return f"Date Range: {start_text} to {end_text}"


def _pdf_cell(value, style):
    if value is None:
        return Paragraph("-", style)

    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            text = value.strftime("%d-%m-%Y %H:%M")
        else:
            text = value.strftime("%d-%m-%Y")
    else:
        text = str(value)

    return Paragraph(
        text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"),
        style,
    )


@router.get("/export/{report_name}/pdf")
def export_report_pdf(
    report_name: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    ),
):
    payload = _report_export_payload(
        report_name=report_name,
        db=db,
        current_user=current_user,
        start_date=start_date,
        end_date=end_date,
    )

    if payload is None:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=404,
            detail="Unsupported report type.",
        )

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=payload["title"],
    )

    styles = getSampleStyleSheet()
    body_style = styles["BodyText"]
    body_style.fontSize = 7
    body_style.leading = 9

    story = [
        Paragraph("Smart Library Management System", styles["Title"]),
        Paragraph(payload["title"], styles["Heading2"]),
        Paragraph(
            "Generated: " + datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
            styles["Normal"],
        ),
    ]

    date_filter_text = _report_date_filter_text(
        start_date,
        end_date,
    )
    if (
        report_name in ("issued-books", "issue-history")
        and date_filter_text
    ):
        story.append(
            Paragraph(date_filter_text, styles["Normal"])
        )

    story.append(Spacer(1, 8))

    if payload.get("summary"):
        summary_rows = [["Summary", "Value"]]
        summary_rows.extend(payload["summary"])
        summary_table = Table(
            summary_rows,
            colWidths=[65 * mm, 45 * mm],
            repeatRows=1,
        )
        summary_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
        )
        story.append(summary_table)
        story.append(Spacer(1, 10))

    table_rows = [
        [_pdf_cell(header, body_style) for header in payload["headers"]]
    ]
    for row in payload["rows"]:
        table_rows.append(
            [_pdf_cell(value, body_style) for value in row]
        )

    available_width = landscape(A4)[0] - (20 * mm)
    column_count = max(1, len(payload["headers"]))
    column_widths = [available_width / column_count] * column_count

    report_table = Table(
        table_rows,
        colWidths=column_widths,
        repeatRows=1,
    )
    report_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    story.append(report_table)

    document.build(story)
    buffer.seek(0)

    filename = (
        f"smart_library_{report_name.replace('-', '_')}_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )


@router.get("/export/{report_name}/excel")
def export_report_excel(
    report_name: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles("ADMIN", "LIBRARIAN")
    ),
):
    payload = _report_export_payload(
        report_name=report_name,
        db=db,
        current_user=current_user,
        start_date=start_date,
        end_date=end_date,
    )

    if payload is None:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=404,
            detail="Unsupported report type.",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    sheet["A1"] = "Smart Library Management System"
    sheet["A1"].font = Font(bold=True, size=16)

    sheet["A2"] = payload["title"]
    sheet["A2"].font = Font(bold=True, size=13)

    sheet["A3"] = "Generated At"
    sheet["B3"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M:%S"
    )

    current_row = 4
    date_filter_text = _report_date_filter_text(
        start_date,
        end_date,
    )
    if (
        report_name in ("issued-books", "issue-history")
        and date_filter_text
    ):
        sheet.cell(current_row, 1, "Date Filter")
        sheet.cell(current_row, 2, date_filter_text)
        current_row += 1

    if payload.get("summary"):
        current_row += 1
        sheet.cell(current_row, 1, "Summary")
        sheet.cell(current_row, 2, "Value")
        for cell in sheet[current_row]:
            if cell.column <= 2:
                cell.fill = header_fill
                cell.font = header_font

        for label, value in payload["summary"]:
            current_row += 1
            sheet.cell(current_row, 1, label)
            sheet.cell(current_row, 2, value)

    current_row += 2

    for column_index, header in enumerate(
        payload["headers"],
        start=1,
    ):
        cell = sheet.cell(
            current_row,
            column_index,
            header,
        )
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    header_row_number = current_row

    for row in payload["rows"]:
        current_row += 1
        for column_index, value in enumerate(
            row,
            start=1,
        ):
            sheet.cell(
                current_row,
                column_index,
                _excel_safe_datetime(value),
            )

    sheet.freeze_panes = f"A{header_row_number + 1}"
    sheet.auto_filter.ref = (
        f"A{header_row_number}:"
        f"{get_column_letter(len(payload['headers']))}{current_row}"
    )

    for column_index, header in enumerate(
        payload["headers"],
        start=1,
    ):
        width = max(14, min(35, len(str(header)) + 8))
        sheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = (
        f"smart_library_{report_name.replace('-', '_')}_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".xlsx"
    )

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )

